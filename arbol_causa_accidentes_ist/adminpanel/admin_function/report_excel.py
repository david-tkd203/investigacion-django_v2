# arbol_causa_accidentes_ist/adminpanel/admin_function/report_excel.py
# -*- coding: utf-8 -*-
from __future__ import annotations
import io
import csv
import datetime
from typing import Optional, Iterable, List, Tuple, Any, Dict

from django.apps import apps
from django.core.paginator import Paginator
from django.views import View
from django.views.generic import TemplateView
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils.timezone import make_naive, is_aware
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.db.models import Exists, OuterRef, Min

from accidentes.models import Accidentes, Informes
from accidentes.access import scope_accidentes_q  # si no existe, se maneja con try/except
from django.contrib.auth import get_user_model
from django.db.models import Q
User = get_user_model()
# ======================= Permisos =======================
ALLOWED_ROLES = {"admin", "admin_ist", "admin_holding", "admin_empresa", "coordinador"}

def _user_can_download(user, accidente: Accidentes | None) -> bool:
    if not getattr(user, "is_authenticated", False):
        return False
    if getattr(user, "rol", None) in ALLOWED_ROLES:
        return True
    if accidente is None:
        return False
    return accidente.usuario_asignado_id == getattr(user, "id", None)

# ======================= Utils generales =======================
def _parse_date(val: str) -> Optional[datetime.date]:
    if not val:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.datetime.strptime(val, fmt).date()
        except Exception:
            pass
    return None

def _to_naive(dt: datetime.datetime | None) -> Optional[datetime.datetime]:
    if not dt:
        return None
    try:
        return make_naive(dt) if is_aware(dt) else dt
    except Exception:
        return dt

def _safe(val, default=""):
    return val if val is not None else default

def _get_model_if_exists(app_label: str, model_name: str):
    try:
        return apps.get_model(app_label, model_name)
    except Exception:
        return None

def _first_attr(obj: Any, candidates: List[str], default: Any = "") -> Any:
    for c in candidates:
        if hasattr(obj, c):
            v = getattr(obj, c)
            if v is not None:
                # si es callable tipo get_full_name
                if callable(v):
                    try:
                        v = v()
                    except Exception:
                        pass
                return v
    return default

def _as_name(x) -> str:
    """Devuelve nombre si x tiene .nombre; si es str, devuelve str; si es obj, str(obj)."""
    if x is None:
        return ""
    if hasattr(x, "nombre"):
        try:
            v = getattr(x, "nombre")
            return str(v or "")
        except Exception:
            pass
    if isinstance(x, str):
        return x
    try:
        return str(x)
    except Exception:
        return ""

def _tenure(trabajador, ref_date: Optional[datetime.date]) -> Tuple[str, str, Optional[datetime.date]]:
    """
    Calcula antigüedad (años, meses) desde fecha_ingreso del trabajador hasta ref_date (fecha_accidente) o hoy.
    Busca múltiples nombres de campo: fecha_ingreso, fecha_contratacion, fecha_inicio, fecha_alta, ingreso, etc.
    """
    if not trabajador:
        return "", "", None
    start: Optional[datetime.date] = None
    start = _first_attr(
        trabajador,
        ["fecha_ingreso", "fecha_contratacion", "fecha_inicio", "fecha_alta", "ingreso"],
        None,
    )
    if isinstance(start, datetime.datetime):
        start = start.date()
    if not isinstance(start, datetime.date):
        start = None

    if not start:
        return "", "", None

    end = ref_date or datetime.date.today()
    if isinstance(end, datetime.datetime):
        end = end.date()

    # cálculo en meses totales
    months = (end.year - start.year) * 12 + (end.month - start.month)
    if end.day < start.day:
        months -= 1  # aproximación
    months = max(0, months)
    years = months // 12
    rem_months = months % 12
    return str(years), str(rem_months), start

def _normalize_contrato(raw) -> str:
    """
    Normaliza cadenas comunes de tipo de contrato.
    """
    s = (str(raw or "")).strip().lower()
    if not s:
        return ""
    # mapeos simples
    if "indef" in s or "ind" in s:
        return "Indefinido"
    if "plazo" in s or "fijo" in s or "determinado" in s:
        return "Plazo fijo"
    if "honorario" in s or "servicio" in s:
        return "Honorarios/Servicios"
    return str(raw)

def _resolve_region_comuna(centro, empresa, trabajador) -> Tuple[str, str]:
    """
    Intenta obtener nombres de Región y Comuna desde centro>empresa>trabajador con varios esquemas:
      - Objeto con .nombre
      - Campo plano 'region'/'comuna' (str)
      - Objetos anidados tipo region.nombre / comuna.nombre
    """
    region = ""
    comuna = ""

    # Preferencia: centro
    if centro:
        region = _as_name(_first_attr(centro, ["region", "region_nombre", "reg"], ""))
        comuna = _as_name(_first_attr(centro, ["comuna", "comuna_nombre", "com"], ""))
    # Fallback: empresa
    if not region and empresa:
        region = _as_name(_first_attr(empresa, ["region", "region_nombre", "reg"], ""))
    if not comuna and empresa:
        comuna = _as_name(_first_attr(empresa, ["comuna", "comuna_nombre", "com"], ""))
    # Fallback: trabajador
    if not region and trabajador:
        region = _as_name(_first_attr(trabajador, ["region", "region_nombre", "reg"], ""))
    if not comuna and trabajador:
        comuna = _as_name(_first_attr(trabajador, ["comuna", "comuna_nombre", "com"], ""))
    return region, comuna

def _domicilio_trabajador(trabajador) -> str:
    return _first_attr(trabajador, ["domicilio", "direccion_particular", "direccion", "domicilio_particular"], "")

def _fecha_nacimiento(trabajador) -> Optional[datetime.date]:
    v = _first_attr(trabajador, ["fecha_nacimiento", "nacimiento", "fec_nac"], None)
    if isinstance(v, datetime.datetime):
        v = v.date()
    if isinstance(v, datetime.date):
        return v
    return None

def _nacionalidad(trabajador) -> str:
    v = _first_attr(trabajador, ["nacionalidad"], "")
    if not v and trabajador:
        # a veces viene como objeto
        v = _as_name(_first_attr(trabajador, ["pais_origen", "country"], ""))
    return str(v or "")

def _estado_civil(trabajador) -> str:
    return _first_attr(trabajador, ["estado_civil", "civil", "estadocivil"], "")

def _tipo_contrato(trabajador) -> str:
    raw = _first_attr(trabajador, ["tipo_contrato", "contrato", "tipoContrato"], "")
    return _normalize_contrato(raw)

# ======================= Query base y límites =======================
def _base_queryset(user):
    qs = (
        Accidentes.objects
        .select_related("empresa", "empresa__holding", "centro", "trabajador", "creado_por", "usuario_asignado")
        .all()
    )
    try:
        q = scope_accidentes_q(user)
        qs = qs.filter(q)
    except Exception:
        if hasattr(Accidentes.objects, "visibles_para"):
            qs = Accidentes.objects.visibles_para(user).select_related(
                "empresa", "empresa__holding", "centro", "trabajador", "creado_por", "usuario_asignado"
            )
    return qs

def _get_date_bounds(user) -> dict:
    qs = _base_queryset(user)
    agg = qs.aggregate(
        min_accidente=Min("fecha_accidente"),
        min_creacion_dt=Min("creado_en"),
    )
    today = datetime.date.today()
    min_accidente = agg.get("min_accidente") or today
    min_creacion_dt = agg.get("min_creacion_dt")
    if isinstance(min_creacion_dt, datetime.datetime):
        min_creacion = min_creacion_dt.date()
    elif isinstance(min_creacion_dt, datetime.date):
        min_creacion = min_creacion_dt
    else:
        min_creacion = today
    return {
        "min_accidente": min_accidente,
        "min_creacion": min_creacion,
        "max_today": today,
    }

def _parse_int(val, default=None):
    try:
        return int(val)
    except (TypeError, ValueError):
        return default


def _filter_role_limits(user, qs):
    """
    Aplica los límites por rol para listas de opciones (no para alcance central).
    - admin/admin_ist: sin límites extra
    - admin_holding: limitar por holding del usuario (si lo tiene)
    - admin_empresa: limitar por empresa del usuario (si la tiene)
    """
    rol = getattr(user, "rol", None)
    if rol in {"admin", "admin_ist"}:
        return qs
    if rol == "admin_holding":
        hid = getattr(user, "holding_id", None)
        if hid:
            qs = qs.filter(empresa__holding_id=hid)
        return qs
    if rol == "admin_empresa":
        eid = getattr(user, "empresa_id", None)
        if eid:
            qs = qs.filter(empresa_id=eid)
        return qs
    return qs.none()


def _get_filter_options(user, *, holding_id=None, empresa_id=None):
    """
    Construye opciones de selects según el rol y selección actual (cascada).
    - admin/admin_ist: holdings, empresas (por holding), inv/coord (por empresa si hay, si no por holding).
    - admin_holding: empresas (por su holding), inv/coord (por empresa si hay, si no por holding).
    - admin_empresa: inv/coord (por su empresa).
    """
    rol = getattr(user, "rol", None)
    base_qs = _base_queryset(user)  # ya viene con scope_accidentes_q

    # NORMALIZAR IDS
    hid = None if holding_id in ("", None) else int(holding_id)
    eid = None if empresa_id in ("", None) else int(empresa_id)

    # ============ HOLDINGS ============
    holdings = []
    if rol in {"admin", "admin_ist"}:
        from accidentes.models import Holdings
        holdings_ids = (
            base_qs.values_list("empresa__holding_id", flat=True)
            .distinct().exclude(empresa__holding_id__isnull=True)
        )
        holdings = Holdings.objects.filter(pk__in=holdings_ids).order_by("nombre")

    # ============ EMPRESAS ============
    from accidentes.models import Empresas
    empresas_all = Empresas.objects.filter(
        pk__in=base_qs.values_list("empresa_id", flat=True).distinct()
    ).order_by("empresa_sel")

    if rol == "admin_holding":
        rol_hid = getattr(user, "holding_id", None)
        empresas = empresas_all.filter(holding_id=rol_hid)
    elif rol == "admin_empresa":
        rol_eid = getattr(user, "empresa_id", None)
        empresas = empresas_all.filter(pk=rol_eid)
    elif rol in {"admin", "admin_ist"} and hid:
        empresas = empresas_all.filter(holding_id=hid)
    else:
        empresas = empresas_all

    # Filtrar base para usuarios según cascada
    # Si hay empresa seleccionada => filtrar por esa empresa
    # Else, si hay holding seleccionado => filtrar por ese holding
    users_base = base_qs
    if eid:
        users_base = users_base.filter(empresa_id=eid)
    elif hid:
        users_base = users_base.filter(empresa__holding_id=hid)
    else:
        # Para admin_empresa, siempre su empresa
        if rol == "admin_empresa":
            rol_eid = getattr(user, "empresa_id", None)
            if rol_eid:
                users_base = users_base.filter(empresa_id=rol_eid)
        # Para admin_holding, su holding
        if rol == "admin_holding":
            rol_hid = getattr(user, "holding_id", None)
            if rol_hid:
                users_base = users_base.filter(empresa__holding_id=rol_hid)

    # ============ INVESTIGADORES ============
    inv_ids = (users_base
               .values_list("usuario_asignado_id", flat=True)
               .distinct()
               .exclude(usuario_asignado_id__isnull=True))
    investigadores = User.objects.filter(pk__in=inv_ids).order_by("first_name", "last_name", "username")

    # ============ COORDINADORES ============
    coord_ids = (users_base
                 .filter(creado_por__rol="coordinador")
                 .values_list("creado_por_id", flat=True)
                 .distinct())
    coordinadores = User.objects.filter(pk__in=coord_ids).order_by("first_name", "last_name", "username")

    return {
        "holdings": holdings,
        "empresas": empresas,
        "investigadores": investigadores,
        "coordinadores": coordinadores,
    }

def _apply_filters(qs, request, user):
    bounds = _get_date_bounds(user)
    date_kind = (request.GET.get("date_kind") or request.POST.get("date_kind") or "accidente").strip()
    d1 = _parse_date(request.GET.get("date_from") or request.POST.get("date_from") or "")
    d2 = _parse_date(request.GET.get("date_to")   or request.POST.get("date_to")   or "")

    field = "creado_en__date" if date_kind == "creacion" else "fecha_accidente"

    min_allowed = bounds["min_creacion"] if date_kind == "creacion" else bounds["min_accidente"]
    max_allowed = bounds["max_today"]

    if d1 and d1 < min_allowed:
        d1 = min_allowed
    if d2 and d2 > max_allowed:
        d2 = max_allowed

    if d1:
        qs = qs.filter(**{f"{field}__gte": d1})
    if d2:
        qs = qs.filter(**{f"{field}__lte": d2})

    # ----------- NUEVOS FILTROS POR ROL -----------
    rol = getattr(user, "rol", None)
    holding_id = _parse_int(request.GET.get("holding_id") or request.POST.get("holding_id"))
    empresa_id = _parse_int(request.GET.get("empresa_id")  or request.POST.get("empresa_id"))
    investigador_id = _parse_int(request.GET.get("investigador_id") or request.POST.get("investigador_id"))
    coordinador_id  = _parse_int(request.GET.get("coordinador_id")  or request.POST.get("coordinador_id"))

    # admin/admin_ist -> puede filtrar por holding
    if rol in {"admin", "admin_ist"} and holding_id:
        qs = qs.filter(empresa__holding_id=holding_id)

    # admin/admin_ist/admin_holding -> puede filtrar por empresa
    if rol in {"admin", "admin_ist", "admin_holding"} and empresa_id:
        qs = qs.filter(empresa_id=empresa_id)
    # admin_empresa: la empresa ya viene dada por su rol; aún así, si envía empresa_id, lo respetamos si coincide.
    if rol == "admin_empresa" and empresa_id:
        qs = qs.filter(empresa_id=empresa_id)

    # investigador (usuario_asignado)
    if investigador_id:
        qs = qs.filter(usuario_asignado_id=investigador_id)

    # coordinador (usamos creado_por con rol 'coordinador')
    if coordinador_id:
        qs = qs.filter(creado_por_id=coordinador_id, creado_por__rol="coordinador")

    # ----------------------------------------------

    has_is_current = any(f.name == "is_current" for f in Informes._meta.fields)
    if has_is_current:
        qs = qs.annotate(
            tiene_informe=Exists(
                Informes.objects.filter(accidente_id=OuterRef("pk"), is_current=True)
            )
        )
    else:
        qs = qs.annotate(
            tiene_informe=Exists(
                Informes.objects.filter(accidente_id=OuterRef("pk"))
            )
        )

    qs = qs.order_by("-fecha_accidente", "-creado_en", "-pk")
    return qs, (d1, d2), date_kind, bounds

# ======================= Helpers de datos detallados =======================
def _current_informe_for_accidente(a: Accidentes):
    has_is_current = any(f.name == "is_current" for f in Informes._meta.fields)
    if has_is_current:
        return Informes.objects.filter(accidente=a, is_current=True).first()
    return Informes.objects.filter(accidente=a).order_by("-version").first()

def _preguntas_for_accidente(a: Accidentes) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    PG = _get_model_if_exists("accidentes", "PreguntasGuia")
    if PG:
        for p in PG.objects.filter(accidente=a).order_by("pk"):
            out.append({
                "tipo": "guia",
                "categoria": _safe(getattr(p, "categoria", "")),
                "pregunta": _safe(getattr(p, "pregunta", "")),
                "objetivo": _safe(getattr(p, "objetivo", "")),
                "respuesta": _safe(getattr(p, "respuesta", "")),
            })
    DEC = _get_model_if_exists("accidentes", "Declaraciones")
    if DEC:
        for d in DEC.objects.filter(accidente=a).order_by("pk"):
            out.append({
                "tipo": "declaracion",
                "categoria": _safe(getattr(d, "tipo_decl", "")),
                "pregunta": _safe(getattr(d, "nombre", "")),
                "objetivo": "",
                "respuesta": _safe(getattr(d, "texto", "")),
            })
    return out

def _relatos_for_accidente(a: Accidentes) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    Relato = _get_model_if_exists("accidentes", "Relato")
    if not Relato:
        return out
    qs = Relato.objects.filter(accidente=a)
    field_names = {f.name for f in Relato._meta.fields}
    if "is_current" in field_names:
        qs = qs.order_by("-is_current", "-pk")
    else:
        qs = qs.order_by("-pk")
    for r in qs:
        out.append({
            "relato_inicial": _safe(getattr(r, "relato_inicial", "")),
            "pregunta_1": _safe(getattr(r, "pregunta_1", "")),
            "respuesta_1": _safe(getattr(r, "respuesta_1", "")),
            "fraseQR1": _safe(getattr(r, "fraseQR1", "")),
            "pregunta_2": _safe(getattr(r, "pregunta_2", "")),
            "respuesta_2": _safe(getattr(r, "respuesta_2", "")),
            "fraseQR2": _safe(getattr(r, "fraseQR2", "")),
            "pregunta_3": _safe(getattr(r, "pregunta_3", "")),
            "respuesta_3": _safe(getattr(r, "respuesta_3", "")),
            "fraseQR3": _safe(getattr(r, "fraseQR3", "")),
            "relato_final": _safe(getattr(r, "relato_final", "")),
            "is_current": _safe(getattr(r, "is_current", "")),
        })
    return out

def _medidas_for_accidente(a: Accidentes) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    Model = (
        _get_model_if_exists("accidentes", "MedidasCorrectivas") or
        _get_model_if_exists("accidentes", "Medidas") or
        _get_model_if_exists("accidentes", "MedidaCorrectiva")
    )
    if not Model:
        return out

    CAND = {
        "descripcion": ["descripcion", "detalle", "texto", "nombre"],
        "responsable": ["responsable", "asignado_a", "encargado"],
        "fecha_compromiso": ["fecha_compromiso", "fecha_plazo", "plazo"],
        "fecha_cierre": ["fecha_cierre", "cerrado_en"],
        "estado": ["estado", "status"],
    }

    for m in Model.objects.filter(accidente=a).order_by("pk"):
        row = {
            "descripcion": _first_attr(m, CAND["descripcion"], ""),
            "responsable": _first_attr(m, CAND["responsable"], ""),
            "fecha_compromiso": _first_attr(m, CAND["fecha_compromiso"], ""),
            "fecha_cierre": _first_attr(m, CAND["fecha_cierre"], ""),
            "estado": _first_attr(m, CAND["estado"], ""),
        }
        out.append(row)
    return out

# ======================= Excel/CSV Builder =======================
def build_excel(rows: Iterable[Accidentes]) -> Tuple[bytes, str]:
    """
    Devuelve (data_bytes, ext) donde ext es 'xlsx' o 'csv' si openpyxl no está disponible.
    Hojas:
      - Casos (con datos ampliados de contrato/persona/ubicación)
      - Preguntas (guía + declaraciones)
      - Relato
      - Medidas
    """
    use_xlsx = True
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment
        from openpyxl.utils import get_column_letter
    except Exception:
        use_xlsx = False

    # ---------- Fallback CSV ----------
    if not use_xlsx:
        out = io.StringIO()
        w = csv.writer(out)
        w.writerow([
            "Código Accidente", "Creado En",
            "Holding", "Empresa", "RUT Empresa",
            "Centro", "Dirección Centro", "Región", "Comuna",
            "Trabajador", "RUT Trabajador", "Domicilio Trabajador",
            "Fecha Nacimiento", "Nacionalidad", "Estado Civil",
            "Tipo Contrato", "Fecha Ingreso", "Antigüedad (años)", "Antigüedad (meses)",
            "Fecha Accidente", "Hora Accidente",
            "Lugar", "Naturaleza Lesión", "Tarea", "Operación",
            "Contexto", "Circunstancias",
            "Investigador (asignado)", "Creador",
            "Tiene Informe", "Código Informe", "Versión Informe", "Fecha Informe",
        ])
        for a in rows:
            emp = getattr(a, "empresa", None)
            cen = getattr(a, "centro", None)
            trab = getattr(a, "trabajador", None)
            current = _current_informe_for_accidente(a)

            creado_en = getattr(a, "creado_en", None)
            if isinstance(creado_en, datetime.datetime):
                creado_en = _to_naive(creado_en)
                creado_en_str = creado_en.strftime("%Y-%m-%d %H:%M:%S")
            else:
                creado_en_str = ""

            region, comuna = _resolve_region_comuna(cen, emp, trab)
            dom_trab = _domicilio_trabajador(trab)
            fnac = _fecha_nacimiento(trab)
            nac = _nacionalidad(trab)
            ecivil = _estado_civil(trab)
            tipo_contrato = _tipo_contrato(trab)
            years, months, f_ing = _tenure(trab, getattr(a, "fecha_accidente", None))

            w.writerow([
                _safe(getattr(a, "codigo_accidente", "")),
                creado_en_str,
                _safe(getattr(getattr(emp, "holding", None), "nombre", "")),
                _safe(getattr(emp, "empresa_sel", "")),
                _safe(getattr(emp, "rut_empresa", "")),
                _safe(getattr(cen, "nombre_local", "")),
                _safe(getattr(cen, "direccion", "")),
                region, comuna,
                _safe(getattr(trab, "nombre_trabajador", "")),
                _safe(getattr(trab, "rut_trabajador", "")),
                dom_trab,
                fnac.strftime("%Y-%m-%d") if isinstance(fnac, datetime.date) else "",
                nac, ecivil,
                tipo_contrato,
                f_ing.strftime("%Y-%m-%d") if isinstance(f_ing, datetime.date) else "",
                years, months,
                _safe(getattr(a, "fecha_accidente", None)) or "",
                _safe(getattr(a, "hora_accidente", None)) or "",
                _safe(getattr(a, "lugar_accidente", "")),
                _safe(getattr(a, "naturaleza_lesion", "")),
                _safe(getattr(a, "tarea", "")),
                _safe(getattr(a, "operacion", "")),
                _safe(getattr(a, "contexto", "")),
                _safe(getattr(a, "circunstancias", "")),
                (getattr(getattr(a, "usuario_asignado", None), "get_full_name", lambda: "")() or
                 _safe(getattr(getattr(a, "usuario_asignado", None), "username", ""))),
                (getattr(getattr(a, "creado_por", None), "get_full_name", lambda: "")() or
                 _safe(getattr(getattr(a, "creado_por", None), "username", ""))),
                "Sí" if current else "No",
                _safe(getattr(current, "codigo", "")),
                _safe(getattr(current, "version", "")),
                _safe(getattr(current, "fecha_informe", None)) or "",
            ])
        return out.getvalue().encode("utf-8-sig"), "csv"

    # ---------- XLSX con openpyxl ----------
    wb = Workbook()

    # ---- Hoja 1: Casos ----
    ws = wb.active
    ws.title = "Casos"

    headers = [
        "Código Accidente", "Creado En",
        "Holding", "Empresa", "RUT Empresa",
        "Centro", "Dirección Centro", "Región", "Comuna",
        "Trabajador", "RUT Trabajador", "Domicilio Trabajador",
        "Fecha Nacimiento", "Nacionalidad", "Estado Civil",
        "Tipo Contrato", "Fecha Ingreso", "Antigüedad (años)", "Antigüedad (meses)",
        "Fecha Accidente", "Hora Accidente",
        "Lugar", "Naturaleza Lesión", "Tarea", "Operación",
        "Contexto", "Circunstancias",
        "Investigador (asignado)", "Creador",
        "Tiene Informe", "Código Informe", "Versión Informe", "Fecha Informe",
    ]
    ws.append(headers)

    idx_fecha_acc = headers.index("Fecha Accidente") + 1
    idx_hora_acc  = headers.index("Hora Accidente") + 1
    idx_creado_en = headers.index("Creado En") + 1
    idx_fecha_inf = headers.index("Fecha Informe") + 1
    idx_fecha_ing = headers.index("Fecha Ingreso") + 1
    idx_fnac      = headers.index("Fecha Nacimiento") + 1

    max_len = [len(h) for h in headers]
    accidentes_list: List[Accidentes] = []

    for a in rows:
        accidentes_list.append(a)
        emp = getattr(a, "empresa", None)
        cen = getattr(a, "centro", None)
        trab = getattr(a, "trabajador", None)
        current = _current_informe_for_accidente(a)

        fecha_acc = getattr(a, "fecha_accidente", None)
        hora_acc  = getattr(a, "hora_accidente", None)

        creado_en = getattr(a, "creado_en", None)
        if isinstance(creado_en, datetime.datetime):
            creado_en = _to_naive(creado_en)

        fecha_inf = getattr(current, "fecha_informe", None)

        # nuevos datos
        region, comuna = _resolve_region_comuna(cen, emp, trab)
        dom_trab = _domicilio_trabajador(trab)
        fnac = _fecha_nacimiento(trab)
        nac = _nacionalidad(trab)
        ecivil = _estado_civil(trab)
        tipo_contrato = _tipo_contrato(trab)
        years, months, f_ing = _tenure(trab, fecha_acc)

        row = [
            _safe(getattr(a, "codigo_accidente", "")),
            creado_en,
            _safe(getattr(getattr(emp, "holding", None), "nombre", "")),
            _safe(getattr(emp, "empresa_sel", "")),
            _safe(getattr(emp, "rut_empresa", "")),
            _safe(getattr(cen, "nombre_local", "")),
            _safe(getattr(cen, "direccion", "")),
            region, comuna,
            _safe(getattr(trab, "nombre_trabajador", "")),
            _safe(getattr(trab, "rut_trabajador", "")),
            dom_trab,
            fnac,
            nac, ecivil,
            tipo_contrato,
            f_ing,
            years, months,
            fecha_acc,
            hora_acc,
            _safe(getattr(a, "lugar_accidente", "")),
            _safe(getattr(a, "naturaleza_lesion", "")),
            _safe(getattr(a, "tarea", "")),
            _safe(getattr(a, "operacion", "")),
            _safe(getattr(a, "contexto", "")),
            _safe(getattr(a, "circunstancias", "")),
            (getattr(getattr(a, "usuario_asignado", None), "get_full_name", lambda: "")() or
             _safe(getattr(getattr(a, "usuario_asignado", None), "username", ""))),
            (getattr(getattr(a, "creado_por", None), "get_full_name", lambda: "")() or
             _safe(getattr(getattr(a, "creado_por", None), "username", ""))),
            "Sí" if current else "No",
            _safe(getattr(current, "codigo", "")),
            _safe(getattr(current, "version", "")),
            fecha_inf,
        ]
        ws.append(row)

        for i, val in enumerate(row):
            if isinstance(val, datetime.datetime):
                s = val.strftime("%Y-%m-%d %H:%M:%S")
            elif isinstance(val, datetime.date):
                s = val.strftime("%Y-%m-%d")
            elif isinstance(val, datetime.time):
                s = val.strftime("%H:%M:%S")
            else:
                s = str(val) if val is not None else ""
            max_len[i] = max(max_len[i], len(s))

    # formatos
    for col in ws.iter_cols(min_col=idx_fecha_acc, max_col=idx_fecha_acc, min_row=2):
        for c in col:
            if isinstance(c.value, (datetime.date, datetime.datetime)):
                c.number_format = "yyyy-mm-dd"
    for col in ws.iter_cols(min_col=idx_fecha_inf, max_col=idx_fecha_inf, min_row=2):
        for c in col:
            if isinstance(c.value, (datetime.date, datetime.datetime)):
                c.number_format = "yyyy-mm-dd"
    for col in ws.iter_cols(min_col=idx_hora_acc, max_col=idx_hora_acc, min_row=2):
        for c in col:
            if isinstance(c.value, datetime.time):
                c.number_format = "HH:mm:ss"
    for col in ws.iter_cols(min_col=idx_creado_en, max_col=idx_creado_en, min_row=2):
        for c in col:
            if isinstance(c.value, datetime.datetime):
                c.number_format = "yyyy-mm-dd HH:mm:ss"
    for col in ws.iter_cols(min_col=idx_fecha_ing, max_col=idx_fecha_ing, min_row=2):
        for c in col:
            if isinstance(c.value, (datetime.date, datetime.datetime)):
                c.number_format = "yyyy-mm-dd"
    for col in ws.iter_cols(min_col=idx_fnac, max_col=idx_fnac, min_row=2):
        for c in col:
            if isinstance(c.value, (datetime.date, datetime.datetime)):
                c.number_format = "yyyy-mm-dd"

    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)

    from openpyxl.utils import get_column_letter
    for idx, length in enumerate(max_len, start=1):
        letter = get_column_letter(idx)
        ws.column_dimensions[letter].width = max(12, min(60, length + 2))

    # ---- Hoja 2: Preguntas ----
    ws2 = wb.create_sheet("Preguntas")
    ws2.append(["Código Accidente", "Tipo", "Categoria", "Pregunta", "Objetivo", "Respuesta"])
    for a in accidentes_list:
        for row in _preguntas_for_accidente(a):
            ws2.append([
                _safe(getattr(a, "codigo_accidente", "")),
                row.get("tipo", ""),
                row.get("categoria", ""),
                row.get("pregunta", ""),
                row.get("objetivo", ""),
                row.get("respuesta", ""),
            ])

    # ---- Hoja 3: Relato ----
    ws3 = wb.create_sheet("Relato")
    ws3.append([
        "Código Accidente",
        "Relato Inicial",
        "Pregunta 1", "Respuesta 1", "FraseQR1",
        "Pregunta 2", "Respuesta 2", "FraseQR2",
        "Pregunta 3", "Respuesta 3", "FraseQR3",
        "Relato Final", "Actual",
    ])
    for a in accidentes_list:
        relatos = _relatos_for_accidente(a)
        if not relatos:
            ws3.append([_safe(getattr(a, "codigo_accidente", ""))] + [""] * 12)
            continue
        for r in relatos:
            ws3.append([
                _safe(getattr(a, "codigo_accidente", "")),
                r.get("relato_inicial", ""),
                r.get("pregunta_1", ""), r.get("respuesta_1", ""), r.get("fraseQR1", ""),
                r.get("pregunta_2", ""), r.get("respuesta_2", ""), r.get("fraseQR2", ""),
                r.get("pregunta_3", ""), r.get("respuesta_3", ""), r.get("fraseQR3", ""),
                r.get("relato_final", ""), r.get("is_current", ""),
            ])

    # ---- Hoja 4: Medidas ----
    ws4 = wb.create_sheet("Medidas")
    ws4.append(["Código Accidente", "Descripción", "Responsable", "Fecha Compromiso", "Fecha Cierre", "Estado"])
    for a in accidentes_list:
        medidas = _medidas_for_accidente(a)
        if not medidas:
            ws4.append([_safe(getattr(a, "codigo_accidente", ""))] + [""] * 5)
            continue
        for m in medidas:
            ws4.append([
                _safe(getattr(a, "codigo_accidente", "")),
                m.get("descripcion", ""),
                m.get("responsable", ""),
                m.get("fecha_compromiso", ""),
                m.get("fecha_cierre", ""),
                m.get("estado", ""),
            ])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), "xlsx"

# ======================= Vistas =======================
class ReporteExcelView(LoginRequiredMixin, TemplateView):
    template_name = "adminpanel/report_excel.html"
    login_url = "/accounts/login/"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["sidebar_active"] = "report_excel"

        # límites al front
        b = _get_date_bounds(self.request.user)
        ctx["date_min_accidente"] = b["min_accidente"].strftime("%Y-%m-%d")
        ctx["date_min_creacion"]  = b["min_creacion"].strftime("%Y-%m-%d")
        ctx["date_max_today"]     = b["max_today"].strftime("%Y-%m-%d")

        # defaults (clamp a límites)
        today = b["max_today"]
        first = (today.replace(day=1) - datetime.timedelta(days=1)).replace(day=1)
        default_from = max(first, b["min_accidente"])  # por defecto “accidente”
        ctx["default_from"] = default_from.strftime("%Y-%m-%d")
        ctx["default_to"]   = today.strftime("%Y-%m-%d")
        # ===== Opciones de filtros por rol =====
        opts = _get_filter_options(self.request.user)
        ctx["filter_holdings"]      = opts["holdings"]
        ctx["filter_empresas"]      = opts["empresas"]
        ctx["filter_investigadores"]= opts["investigadores"]
        ctx["filter_coordinadores"] = opts["coordinadores"]

        # valores seleccionados (para mantener estado)
        ctx["sel_holding_id"]      = self.request.GET.get("holding_id") or self.request.POST.get("holding_id") or ""
        ctx["sel_empresa_id"]      = self.request.GET.get("empresa_id") or self.request.POST.get("empresa_id") or ""
        ctx["sel_investigador_id"] = self.request.GET.get("investigador_id") or self.request.POST.get("investigador_id") or ""
        ctx["sel_coordinador_id"]  = self.request.GET.get("coordinador_id") or self.request.POST.get("coordinador_id") or ""

        ctx["user_rol"] = getattr(self.request.user, "rol", None)
        return ctx

    def post(self, request):
        qs = _base_queryset(request.user)
        qs, (d1, d2), date_kind, bounds = _apply_filters(qs, request, request.user)

        count = qs.count()
        if count == 0:
            messages.info(request, "No hay casos para exportar con los filtros seleccionados.")
            return self.render_to_response({
                **self.get_context_data(),
                "date_from": request.POST.get("date_from", ""),
                "date_to": request.POST.get("date_to", ""),
                "date_kind": date_kind,
            })

        if getattr(request.user, "rol", None) not in ALLOWED_ROLES:
            qs = qs.filter(usuario_asignado_id=getattr(request.user, "id", None))

        data, ext = build_excel(qs.iterator())

        d1s = d1.strftime("%Y-%m-%d") if d1 else "sin_desde"
        d2s = d2.strftime("%Y-%m-%d") if d2 else "sin_hasta"
        filename = f"reporte_casos_{date_kind}_{d1s}_a_{d2s}.{ext}"

        content_type = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            if ext == "xlsx" else
            "text/csv; charset=utf-8"
        )
        resp = HttpResponse(data, content_type=content_type)
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp


class ReporteExcelPreviewHTMX(LoginRequiredMixin, View):
    login_url = "/accounts/login/"

    def get(self, request):
        qs = _base_queryset(request.user)
        qs, (d1, d2), date_kind, bounds = _apply_filters(qs, request, request.user)
        count = qs.count()

        html = render_to_string(
            "adminpanel/partials/report/_preview.html",
            {
                "count": count,
                "date_from": request.GET.get("date_from", ""),
                "date_to": request.GET.get("date_to", ""),
                "date_kind": date_kind,
            },
            request=request,
        )
        return HttpResponse(html, status=200)


class ReporteExcelTableHTMX(LoginRequiredMixin, View):
    """
    Devuelve la tabla HTML de los casos según filtros, con paginación.
    Respeta el mismo alcance que el Excel:
      - Si el rol del usuario está en ALLOWED_ROLES -> ve todos en su alcance.
      - Si no, se filtra a accidentes donde es usuario_asignado.
    """
    login_url = "/accounts/login/"

    def get(self, request):
        qs = _base_queryset(request.user)
        qs, (d1, d2), date_kind, bounds = _apply_filters(qs, request, request.user)

        if getattr(request.user, "rol", None) not in ALLOWED_ROLES:
            qs = qs.filter(usuario_asignado_id=getattr(request.user, "id", None))

        try:
            page = int(request.GET.get("page") or 1)
        except ValueError:
            page = 1
        try:
            page_size = int(request.GET.get("page_size") or 25)
        except ValueError:
            page_size = 25
        page_size = max(5, min(page_size, 200))

        paginator = Paginator(qs, page_size)
        page_obj = paginator.get_page(page)

        html = render_to_string(
            "adminpanel/partials/report/_table.html",
            {
                "page_obj": page_obj,
                "paginator": paginator,
                "date_from": request.GET.get("date_from", ""),
                "date_to": request.GET.get("date_to", ""),
                "date_kind": request.GET.get("date_kind", "accidente"),
                "page_size": page_size,
            },
            request=request,
        )
        return HttpResponse(html, status=200)


class ReporteExcelFiltersHTMX(LoginRequiredMixin, View):
    """
    Re-renderiza la fila de filtros en cascada.
    Además devuelve el 'preview' con hx-swap-oob para actualizar el conteo.
    """
    login_url = "/accounts/login/"

    def get(self, request):
        holding_id     = request.GET.get("holding_id")
        empresa_id     = request.GET.get("empresa_id")
        investigador_id= request.GET.get("investigador_id")
        coordinador_id = request.GET.get("coordinador_id")

        # Opciones en cascada:
        opts = _get_filter_options(
            request.user,
            holding_id=holding_id,
            empresa_id=empresa_id,
        )

        # También recalculamos el preview (conteo) con estos filtros
        qs = _base_queryset(request.user)
        # Inyectamos en request GET los parámetros que ya están llegando; _apply_filters se encarga
        qs, (d1, d2), date_kind, _ = _apply_filters(qs, request, request.user)
        count = qs.count()

        html_filters = render_to_string(
            "adminpanel/partials/report/_filters.html",
            {
                "user_rol": getattr(request.user, "rol", None),
                "filter_holdings": opts["holdings"],
                "filter_empresas": opts["empresas"],
                "filter_investigadores": opts["investigadores"],
                "filter_coordinadores": opts["coordinadores"],
                "sel_holding_id": holding_id or "",
                "sel_empresa_id": empresa_id or "",
                "sel_investigador_id": investigador_id or "",
                "sel_coordinador_id":  coordinador_id or "",
            },
            request=request,
        )

        # Preview OOB
        html_preview = render_to_string(
            "adminpanel/partials/report/_preview.html",
            {
                "count": count,
                "date_from": request.GET.get("date_from", ""),
                "date_to": request.GET.get("date_to", ""),
                "date_kind": request.GET.get("date_kind", "accidente"),
            },
            request=request,
        )
        html_preview_oob = f'<div id="preview-box" hx-swap-oob="true">{html_preview}</div>'

        return HttpResponse(html_filters + html_preview_oob, status=200)